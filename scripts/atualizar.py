#!/usr/bin/env python3
"""
SEED BASE — atualização automática do dashboard.

Baixa as duas planilhas do Google Drive, processa e injeta os dados
no index.html. Roda dentro do GitHub Actions.

Configuração via variáveis de ambiente (secrets do repositório):
  ID_ANALISES  → ID do arquivo ANALISES_DE_SEMENTES_DE_VERAO_25-26.xlsx
  ID_BENEF     → ID do arquivo LOTES_BENEFICIADOS_VERAO_25-26.xlsx
"""

import os
import re
import io
import json
import sys
from datetime import datetime, timezone, timedelta
from collections import defaultdict

import requests
import pandas as pd
import openpyxl

# ---------------------------------------------------------------- config

HTML_PATH = "index.html"

ID_ANALISES = os.environ.get("ID_ANALISES", "").strip()
ID_BENEF    = os.environ.get("ID_BENEF", "").strip()

# Fuso de Brasília — o GitHub Actions roda em UTC
HOJE = datetime.now(timezone(timedelta(hours=-3))).strftime("%d/%m/%Y")

# Carimbo com data e hora (Brasília, UTC-3 fixo) — usado apenas no cabeçalho "ATUALIZAÇÃO"
CARIMBO_HEADER = datetime.now(timezone(timedelta(hours=-3))).strftime("%d/%m/%Y %H:%M")

APROVADOS = {"A", "B", "BV"}

# Nomes usados por terceiros que não batem com o código oficial do CV_META —
# a planilha LOTES_BENEFICIADOS mistura, dependendo da UBS/obtentor, ora o
# nome fantasia (cultivares Brasmax) ora o código com pequenas variações de
# espaço/sufixo (cultivares Corteva). Sem essa normalização, o mesmo cultivar
# aparecia como "cultivares" diferentes na aba Beneficiamento (ex.: ZEUS de
# um lado e 55I57RSF IPRO do outro), fragmentando os totais.
BENEF_ALIASES = {
    "VALENTE": "6968RSF RR",
    "C 2605 E": "C2605 E",
    "TITANIUM": "56IX58RSF I2X",
    "ZEUS": "55I57RSF IPRO",
    "ORION": "55IX56RSF I2X",
    "RAÇA": "63E66RSF E",
    "IMUNE": "53IX55RSF I2X",
    "VÊNUS": "57K58RSF CE",
    "GRAFENO": "57K56RSF CE",
    "RAIO": "50I52RSF IPRO",
    "FIBRA": "64I61RSF IPRO",
    "BATALHA": "68K66RSF CE",
    "NEXUS": "64IX66RSF I2X",
    "COMPACTA": "65I65RSF IPRO",
    # variações de grafia do mesmo código (espaço a mais/faltando, sufixo omitido)
    "C2615": "C2615CE",
    "C2615 CE": "C2615CE",
    "64I61 RSF IPRO": "64I61RSF IPRO",
    "C2645 CE": "C2645CE",
    "C2645": "C2645CE",
}
_BENEF_ALIASES_UP = {k.upper(): v for k, v in BENEF_ALIASES.items()}


def log(msg):
    print(msg, flush=True)


# ---------------------------------------------------------------- download

def extrair_id(valor):
    """Aceita o ID puro ou uma URL completa colada no secret."""
    v = (valor or "").strip().strip('"').strip("'")
    if not v:
        return ""
    # .../spreadsheets/d/ID/edit  ou  .../file/d/ID/view
    m = re.search(r"/d/([a-zA-Z0-9_-]{20,})", v)
    if m:
        return m.group(1)
    # ...?id=ID  ou  ...&id=ID
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]{20,})", v)
    if m:
        return m.group(1)
    return v


def baixar(file_id, nome, obrigatorio=True):
    """Baixa um arquivo do Drive. Funciona para .xlsx e para Sheets nativo.

    Se obrigatorio=False e o ID não estiver configurado, devolve None em vez
    de abortar — o chamador decide o que fazer.
    """
    file_id = extrair_id(file_id)

    if not file_id:
        if not obrigatorio:
            log(f"  {nome}: ID não configurado — etapa será pulada")
            return None
        raise SystemExit(f"ERRO: ID do arquivo '{nome}' não configurado nos secrets.")

    log(f"  {nome}: id={file_id[:8]}...{file_id[-4:]} ({len(file_id)} caracteres)")

    tentativas = [
        ("Sheets nativo",
         f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"),
        ("Drive usercontent",
         f"https://drive.usercontent.google.com/download?id={file_id}&export=download"),
        ("Drive uc (legado)",
         f"https://drive.google.com/uc?export=download&id={file_id}"),
    ]

    sessao = requests.Session()
    sessao.headers.update({"User-Agent": "Mozilla/5.0 (compatible; SeedBaseBot/1.0)"})

    for rotulo, url in tentativas:
        try:
            r = sessao.get(url, timeout=120, allow_redirects=True)
        except Exception as e:
            log(f"    [{rotulo}] falhou na conexão: {type(e).__name__}: {e}")
            continue

        ctype = r.headers.get("content-type", "?").split(";")[0]
        log(f"    [{rotulo}] HTTP {r.status_code} · {ctype} · {len(r.content):,} bytes")

        if r.status_code != 200:
            continue

        if r.content[:2] == b"PK":          # assinatura de xlsx/zip
            log(f"    -> OK via {rotulo}")
            return io.BytesIO(r.content)

        # Arquivo grande: Google devolve uma página pedindo confirmação
        if b"confirm" in r.content[:60000].lower():
            token = re.search(rb'name="confirm"\s+value="([^"]+)"', r.content)
            uuid  = re.search(rb'name="uuid"\s+value="([^"]+)"', r.content)
            if token:
                params = {"id": file_id, "export": "download",
                          "confirm": token.group(1).decode()}
                if uuid:
                    params["uuid"] = uuid.group(1).decode()
                r2 = sessao.get("https://drive.usercontent.google.com/download",
                                params=params, timeout=180)
                if r2.status_code == 200 and r2.content[:2] == b"PK":
                    log(f"    -> OK via {rotulo} (após confirmação)")
                    return io.BytesIO(r2.content)

        # Diagnóstico: mostrar o começo do que veio
        amostra = r.content[:300].decode("utf-8", errors="replace").replace("\n", " ")
        log(f"    conteúdo recebido (início): {amostra[:200]}")

    raise SystemExit(
        f"\nERRO: não consegui baixar '{nome}' (id={file_id}).\n"
        f"Verifique, nesta ordem:\n"
        f"  1. O arquivo está compartilhado como 'Qualquer pessoa com o link' "
        f"com permissão de Leitor?\n"
        f"  2. O ID no secret corresponde a ESTE arquivo?\n"
        f"  3. O arquivo está na sua conta pessoal ou em um Drive corporativo "
        f"com restrição de domínio? Contas corporativas costumam bloquear "
        f"acesso externo mesmo com o link ativo.\n"
        f"Os códigos HTTP acima indicam o motivo: 404 = ID errado, "
        f"403 = sem permissão pública, 200 com HTML = página de login."
    )


# ---------------------------------------------------------------- helpers

def nz(v):
    if isinstance(v, str):
        return None if v.strip() == "" else v
    return None if pd.isna(v) else v


def r1(v):
    if v is None:
        return None
    try:
        return round(float(str(v).strip()), 1)
    except (TypeError, ValueError):
        return None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("="):
        return None  # fórmula não avaliada
    try:
        return float(str(v).replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------- classificação

def classificar(germ, vigor, germ_of, vigor_of, tz_rank):
    def gb(g):
        if g is None:
            return None
        return "A" if g >= 95 else "B" if g >= 85 else "C" if g >= 80 else "D"

    def vb(v):
        if v is None:
            return None
        return "A" if v >= 95 else "B" if v >= 80 else "BV" if v >= 70 else "D"

    rg, rv = gb(germ), vb(vigor)

    if germ_of is not None and vigor_of is not None:
        if vigor_of < 80 or germ_of < 80:
            final = "D"
        elif germ_of >= 95:
            final = "A"
        elif germ_of >= 90:
            final = "B"
        else:
            final = "C"
    elif vigor_of is not None:
        final = "A" if vigor_of >= 95 else "B" if vigor_of >= 80 else "D"
    else:
        if rg is None and rv is None:
            final = tz_rank if tz_rank else "X"
        elif rg == "D":
            final = "D"
        elif rg == "C":
            final = "D" if rv in ("BV", "D") else "C"
        elif rg in ("B", "A"):
            if rv == "D":
                final = "D"
            elif rv == "BV":
                final = "D" if tz_rank == "C" else "BV"
            elif rv == "B":
                final = "C" if tz_rank == "C" else "B"
            else:
                final = {"A": "B", "B": "C"}[rg] if tz_rank == "C" else rg
        else:
            final = "X"

    rotulos = {
        "A": "Ótimo", "B": "Bom", "C": "Condicional",
        "BV": "Condicional Vigor", "D": "Ruim", "X": "Pendente",
    }
    return rg, rv, final, rotulos[final]


# ---------------------------------------------------------------- ANALISES

def parse_analises(buf):
    df = pd.read_excel(buf, sheet_name="SOJA 25-26")
    df = df[df["LOTE"].notna()].reset_index(drop=True)

    # Coluna de hipoclorito (DM %) — resolvida por busca aproximada no
    # cabeçalho para tolerar pequenas variações de espaço/caixa no nome
    # exato da planilha, em vez de travar o pipeline com um KeyError.
    col_hipoclorito = next(
        (c for c in df.columns if "HIPOCLORITO" in str(c).strip().upper()), None
    )

    lots = []
    for _, row in df.iterrows():
        germ     = r1(nz(row["GER. ROLO DE PAPEL %"]))
        vigor    = r1(nz(row["ROLO DE PAPEL E.A. %"]))
        germ_of  = r1(nz(row["GERM OFICIAL %"]))
        vigor_of = r1(nz(row["VIGOR OFICIAL %"]))
        tz_rank  = nz(row["RAKING TZ GERM"])
        rg, rv, ranking, status = classificar(germ, vigor, germ_of, vigor_of, tz_rank)

        kg_raw = nz(row["QUANT.    KG"])
        umid   = nz(row["UMIDADE %"])
        hipoclorito = r1(nz(row[col_hipoclorito])) if col_hipoclorito else None

        lots.append({
            "cultivar": str(nz(row["CULTIVAR"])).strip() if nz(row["CULTIVAR"]) else None,
            "lote": nz(row["LOTE"]),
            "peneira": nz(row["PENEIRA "]),
            "categoria": nz(row["CATEGORIA "]),
            "ubs": nz(row["UBS"]),
            "umidade": round(float(umid), 2) if umid is not None else None,
            "hipoclorito": hipoclorito,
            "tz_viab": r1(nz(row["TZ VIABILDIADE (1-5) "])),
            "tz_vigor": r1(nz(row["TZ VIGOR (1-3)"])),
            "rank_tz": tz_rank,
            "germ": germ, "rank_germ": rg,
            "vigor": vigor, "rank_vigor": rv,
            "germ_oficial": germ_of, "vigor_oficial": vigor_of,
            "modo_classificacao": (
                "ambos" if (germ_of and vigor_of) else ("vigor" if vigor_of else "nenhum")
            ),
            "ranking": ranking, "status": status,
            "bags": int(row["QUANT BB"]) if pd.notna(row["QUANT BB"]) else None,
            "kg": int(kg_raw) if kg_raw and str(kg_raw).strip() not in ("", "None") else None,
            "pms_etq": r1(nz(row["PMS (g) ETIQUETA"])),
            "peso_bag": r1(nz(row["PESO BAG"])),
            "sementes_m": r1(nz(row["N° SEMENTES (MILHÕES)"])),
        })

    return lots


def montar_summary(lots):
    df = pd.DataFrame(lots)
    df["germ_eff"]  = df["germ_oficial"].combine_first(df["germ"])
    df["vigor_eff"] = df["vigor_oficial"].combine_first(df["vigor"])

    summary = []
    for cv, grp in df.groupby("cultivar"):
        ap = grp[grp["ranking"].isin(APROVADOS)]
        c = grp["ranking"].value_counts().to_dict()
        summary.append({
            "cultivar": cv,
            "total_kg": int(grp["kg"].sum()),
            "lotes": int(len(grp)),
            "germ_media": r1(ap["germ_eff"].dropna().mean()),
            "vigor_media": r1(ap["vigor_eff"].dropna().mean()),
            "tz_viab_media": r1(grp["tz_viab"].dropna().mean()),
            "tz_vigor_media": r1(grp["tz_vigor"].dropna().mean()),
            "rank_A": int(c.get("A", 0)), "rank_B": int(c.get("B", 0)),
            "rank_BV": int(c.get("BV", 0)), "rank_C": int(c.get("C", 0)),
            "rank_D": int(c.get("D", 0)), "rank_X": int(c.get("X", 0)),
        })
    summary.sort(key=lambda s: -s["total_kg"])

    pms_cv = []
    for cv, grp in df.groupby("cultivar"):
        e = {"cultivar": cv}
        total_bags = 0
        for pen, suf in [("P1", "p1"), ("P2", "p2")]:
            sub = grp[grp["peneira"] == pen]
            bags = int(sub["bags"].sum()) if len(sub) else 0
            total_bags += bags
            ap = sub[sub["ranking"].isin(APROVADOS)]
            c = sub["ranking"].value_counts().to_dict()
            e[f"bags_{suf}"]   = bags
            e[f"pms_{suf}"]    = r1(ap["pms_etq"].dropna().mean())
            e[f"aprov_{suf}"]  = int(c.get("A", 0) + c.get("B", 0) + c.get("BV", 0))
            e[f"reprov_{suf}"] = int(c.get("C", 0) + c.get("D", 0))
            e[f"pend_{suf}"]   = int(c.get("X", 0))
        e["bags_total"] = total_bags
        pms_cv.append(e)
    pms_cv.sort(key=lambda p: -p["bags_total"])

    return summary, pms_cv


# ---------------------------------------------------------------- DIAGNÓSTICO DA SAFRA
#
# Gera o texto narrativo da aba "Diagnóstico da Safra" (DIAG_DATA) a partir dos
# dados reais desta execução, em vez de deixá-lo estático. Antes, esse texto
# era escrito uma vez à mão e nunca mais tocado pela automação — foi encontrado
# citando "379 lotes... 42.7% de aprovação" numa safra que já ia em 708 lotes e
# 66%, ou seja, desatualizado há semanas. As faixas de status (crítico/regular/
# bom/excelente) e os textos-modelo abaixo foram reconstruídos a partir do
# conteúdo antigo do DIAG_DATA; ajuste aqui se o critério de negócio for outro.

def _status_cultivar(pct):
    if pct >= 99.95:
        return (
            "excelente",
            "Melhor desempenho da safra — todos os lotes aprovados.",
            "Cultivar recomendado para comercialização sem restrições.",
        )
    if pct >= 80:
        return (
            "bom",
            "Bom desempenho geral.",
            "Comercialização possível com verificação individual dos lotes condicionais e reprovados.",
        )
    if pct >= 50:
        return (
            "regular",
            "Desempenho intermediário, requer atenção.",
            "Avaliar individualmente cada lote antes da comercialização.",
        )
    return (
        "critico",
        "Desempenho crítico — maioria dos lotes reprovada.",
        "Uso comercial restrito — verificar lotes individualmente e avaliar destino dos reprovados.",
    )


def _diagnostico_cultivar(row, lots_cv):
    total = row["lotes"]
    aprov = row["rank_A"] + row["rank_B"] + row["rank_BV"]
    cond, reprov, pend = row["rank_C"], row["rank_D"], row["rank_X"]
    pct = round(aprov / total * 100, 1) if total else 0.0

    partes = [f"{total} lotes analisados, totalizando {round(row['total_kg'] / 1000)}t."]

    detalhes = [f"{aprov} aprovados"]
    if cond:
        detalhes.append(f"{cond} condicionais")
    if reprov:
        detalhes.append(f"{reprov} reprovados")
    if pend:
        detalhes.append(f"{pend} pendentes")
    partes.append(f"Taxa de aprovação de {pct}% — {', '.join(detalhes)}.")

    if row["germ_media"] is not None and row["vigor_media"] is not None:
        partes.append(f"Germinação média de {row['germ_media']}% e vigor médio de {row['vigor_media']}%.")

    if reprov:
        vigor_fail = sum(
            1 for l in lots_cv
            if l["ranking"] == "D" and l.get("vigor_oficial") is not None and l["vigor_oficial"] < 80
        )
        germ_fail = sum(
            1 for l in lots_cv
            if l["ranking"] == "D" and l.get("germ_oficial") is not None and l["germ_oficial"] < 80
        )
        if vigor_fail or germ_fail:
            if vigor_fail >= germ_fail:
                partes.append(f"Principal causa de reprovação: vigor abaixo de 80% em {vigor_fail} lotes.")
            else:
                partes.append(f"Principal causa de reprovação: germinação abaixo de 80% em {germ_fail} lotes.")

    if pend:
        bags_pend = sum((l.get("bags") or 0) for l in lots_cv if l["ranking"] == "X")
        partes.append(f"{pend} lotes com {bags_pend} bags ainda sem análise completa.")

    return " ".join(partes), pct, aprov, cond, reprov, pend, total


def montar_diagnostico(lots, summary):
    df = pd.DataFrame(lots)
    df["germ_eff"]  = df["germ_oficial"].combine_first(df["germ"])
    df["vigor_eff"] = df["vigor_oficial"].combine_first(df["vigor"])

    total   = len(df)
    n_cvs   = df["cultivar"].nunique()
    aprov   = int(df["ranking"].isin(APROVADOS).sum())
    cond    = int((df["ranking"] == "C").sum())
    reprov  = int((df["ranking"] == "D").sum())
    pend    = int((df["ranking"] == "X").sum())
    pct     = round(aprov / total * 100, 1) if total else 0.0

    ap = df[df["ranking"].isin(APROVADOS)]
    germ_g  = r1(ap["germ_eff"].dropna().mean())
    vigor_g = r1(ap["vigor_eff"].dropna().mean())
    tz_g    = r1(df["tz_viab"].dropna().mean())

    geral = (
        f"A safra conta com {total} lotes analisados, distribuídos em {n_cvs} cultivares. "
        f"A taxa de aprovação está em {pct}%, com {aprov} lotes aprovados (A+B), {cond} condicionais (C), "
        f"{reprov} reprovados (D) e {pend} aguardando análise (X). "
        f"A média geral de germinação é {germ_g}%, vigor {vigor_g}% e TZ viabilidade {tz_g}%."
    )

    d_lots = df[df["ranking"] == "D"]
    vigor_fail_g = int((d_lots["vigor_oficial"] < 80).sum())
    germ_fail_g  = int((d_lots["germ_oficial"] < 80).sum())
    if reprov == 0:
        problema = "Nenhum lote foi reprovado na safra até o momento."
    elif vigor_fail_g >= germ_fail_g:
        problema = (
            f"O principal critério responsável pelas reprovações é o Vigor EA. "
            f"{vigor_fail_g} lotes foram reprovados com vigor abaixo de 80%, representando a maioria dos "
            f"{reprov} lotes classificados como Ruim (D). A germinação abaixo de 80% contribui com outros "
            f"{germ_fail_g} lotes reprovados. Como o sistema utiliza o pior critério entre os três, qualquer "
            f"valor abaixo do mínimo resulta em reprovação do lote inteiro."
        )
    else:
        problema = (
            f"O principal critério responsável pelas reprovações é a Germinação. "
            f"{germ_fail_g} lotes foram reprovados com germinação abaixo de 80%, representando a maioria dos "
            f"{reprov} lotes classificados como Ruim (D). O vigor abaixo de 80% contribui com outros "
            f"{vigor_fail_g} lotes reprovados. Como o sistema utiliza o pior critério entre os três, qualquer "
            f"valor abaixo do mínimo resulta em reprovação do lote inteiro."
        )

    if pend == 0:
        pendentes_txt = "Não há lotes pendentes de análise no momento."
    else:
        bags_pend_g = int(df.loc[df["ranking"] == "X", "bags"].fillna(0).sum())
        pendentes_txt = (
            f"{pend} lotes ainda não possuem análise de vigor e germinação, totalizando {bags_pend_g} bags "
            f"sem classificação definida. Estes lotes permanecem como pendentes até a conclusão das análises "
            f"laboratoriais."
        )

    cultivares_out = []
    for row in summary:
        lots_cv = [l for l in lots if l["cultivar"] == row["cultivar"]]
        diagnostico, pct_c, aprov_c, cond_c, reprov_c, pend_c, total_c = _diagnostico_cultivar(row, lots_cv)
        cor, status_txt, recomendacao = _status_cultivar(pct_c)
        cultivares_out.append({
            "cultivar": row["cultivar"],
            "volume": f"{round(row['total_kg'] / 1000)}t",
            "pct_aprov": pct_c,
            "status_cor": cor,
            "status_txt": status_txt,
            "diagnostico": diagnostico,
            "recomendacao": recomendacao,
            "aprovados": aprov_c, "condicionais": cond_c, "reprovados": reprov_c,
            "pendentes": pend_c, "total": total_c,
        })

    return {
        "global": {"geral": geral, "problema": problema, "pendentes": pendentes_txt},
        "cultivares": cultivares_out,
    }


def parse_comprados(buf):
    wb = openpyxl.load_workbook(buf)
    ws = wb["LOTES COMPRADOS"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    def _norm(s): return str(s).strip().upper()
    col_norm = {_norm(h): i + 1 for i, h in enumerate(header) if h}

    def cell(r, nome):
        idx = col.get(nome)
        if idx is None:
            idx = col_norm.get(_norm(nome))
        return ws.cell(r, idx).value if idx else None

    def txt(r, nome):
        v = cell(r, nome)
        return str(v).strip() if v else None

    out = []
    for r in range(2, ws.max_row + 1):
        cultivar = ws.cell(r, 1).value
        if not cultivar:
            continue

        bags     = to_float(cell(r, "QUANT BB"))
        peso_bag = to_float(cell(r, "PESO BAG"))
        kg_raw   = to_float(cell(r, "QUANT.    KG"))
        kg = int(bags * peso_bag) if bags and peso_bag else (int(kg_raw) if kg_raw else None)

        # PMS pode vir como fórmula (=PESO BAG/5) — nesse caso calcula na mão
        pms_raw = cell(r, "PMS (g) ETIQUETA")
        if pms_raw is not None and not (isinstance(pms_raw, str) and pms_raw.startswith("=")):
            pms_etq = round(float(pms_raw), 1)
        elif peso_bag:
            pms_etq = round(peso_bag / 5, 1)
        else:
            pms_etq = None

        out.append({
            "cultivar": str(cultivar).strip(),
            "lote": txt(r, "LOTE SCV"),
            "lote_origem": txt(r, "LOTE DE ORIGEM"),
            "empresa": txt(r, "EMPRESA"),
            "germ_oficial": to_float(cell(r, "GERM OFICIAL %")),
            "vigor_oficial": to_float(cell(r, "VIGOR OFICIAL %")),
            "bags": int(bags) if bags else None,
            "peso_bag": peso_bag,
            "kg": kg,
            "categoria_origem": str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else None,
            "categoria_reembalado": str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else None,
            "peneira": txt(r, "PENEIRA "),
            "pms_etq": pms_etq,
            "status": txt(r, "STATUS"),
        })
    return out


# ---------------------------------------------------------------- BENEFICIADOS

# Cada aba tem layout próprio. (linha_inicial, col_cultivar, col_categoria,
# col_peneira, col_pms, col_bags, pms_vem_do_peso)
LAYOUT_BENEF = {
    "SOJA 25-26": ("SCV",      3, 2, 4, 6, 7, 9, False),
    "CAMBAÍ":     ("CAMBAÍ",   3, 1, 3, 5, 9, 8, True),
    "SIMÃO":      ("SIMÃO",    4, 1, 3, 5, 6, 8, False),
    "GIOVELLI":   ("GIOVELLI", 3, 1, 3, 5, 6, 8, False),
    "CERENNA":    ("CERENNA",  4, 1, 3, 5, 6, 8, False),
}

CV_INVALIDOS = {
    "CULTIVAR", "DATA", "SAFRA", "CAT.", "N° BB", "QUANTIDADE",
    "LOTE", "PEN.", "PMS", "EMBALAGEM", "PESO", "LOCALIZAÇÃO",
    "STATUS", "OBSERVAÇÕES",
}


def parse_benef(buf):
    wb = openpyxl.load_workbook(buf)
    agrupado = defaultdict(lambda: {"bags": 0, "kg": 0.0, "pms_sum": 0.0, "pms_cnt": 0})

    for aba, cfg in LAYOUT_BENEF.items():
        if aba not in wb.sheetnames:
            log(f"  aviso: aba '{aba}' não encontrada — ignorando")
            continue

        ubs, linha0, c_cv, c_cat, c_pen, c_pms, c_bags, pms_do_peso = cfg
        ws = wb[aba]

        for r in range(linha0, ws.max_row + 1):
            cv = ws.cell(r, c_cv).value
            if not cv:
                continue
            cv = str(cv).strip()
            if cv in CV_INVALIDOS or "PARA SEMENTES" in cv or "SEMENTES COM VIGOR" in cv:
                continue
            cv = _BENEF_ALIASES_UP.get(cv.upper(), cv)

            bags = to_float(ws.cell(r, c_bags).value)
            if not bags or bags <= 0:
                continue
            bags = int(bags)

            bruto = to_float(ws.cell(r, c_pms).value)
            if not bruto:
                continue
            pms = bruto / 5 if pms_do_peso else bruto  # coluna traz peso do bag

            pen_raw = ws.cell(r, c_pen).value
            pen = "P1" if pen_raw and "P1" in str(pen_raw).upper() else "P2"
            cat_raw = ws.cell(r, c_cat).value
            cat = str(cat_raw).strip() if cat_raw else "C2"

            k = agrupado[(cv, ubs, cat, pen)]
            k["bags"]    += bags
            k["kg"]      += bags * pms * 5
            k["pms_sum"] += pms * bags
            k["pms_cnt"] += bags

    # consolidar por (cultivar, ubs)
    por_cv_ubs = defaultdict(lambda: defaultdict(
        lambda: {"bags": 0, "kg": 0.0, "pms_sum": 0.0, "pms_cnt": 0}
    ))
    for (cv, ubs, cat, pen), v in agrupado.items():
        d = por_cv_ubs[(cv, ubs)][(cat, pen)]
        for campo in ("bags", "kg", "pms_sum", "pms_cnt"):
            d[campo] += v[campo]

    benef = []
    for (cv, ubs), detalhes in por_cv_ubs.items():
        detail = []
        p1_s = p1_n = p2_s = p2_n = 0
        for (cat, pen), v in sorted(detalhes.items()):
            if v["bags"] == 0:
                continue
            detail.append({
                "categoria": cat, "peneira": pen,
                "bags": v["bags"], "kg": round(v["kg"]),
            })
            if pen == "P1":
                p1_s += v["pms_sum"]; p1_n += v["pms_cnt"]
            else:
                p2_s += v["pms_sum"]; p2_n += v["pms_cnt"]

        total = sum(d["bags"] for d in detail)
        if total == 0:
            continue

        benef.append({
            "cultivar": cv, "ubs": ubs,
            "bags_total": total,
            "kg_total": sum(d["kg"] for d in detail),
            "bags_p1": sum(d["bags"] for d in detail if d["peneira"] == "P1"),
            "bags_p2": sum(d["bags"] for d in detail if d["peneira"] == "P2"),
            "pms_todos_p1": round(p1_s / p1_n, 1) if p1_n else None,
            "pms_todos_p2": round(p2_s / p2_n, 1) if p2_n else None,
            "detail": detail,
        })

    benef.sort(key=lambda x: -x["kg_total"])
    return benef


# ---------------------------------------------------------------- injeção

def js(dados):
    return json.dumps(dados, ensure_ascii=False, separators=(",", ":"))


def substituir_const(html, nome, dados):
    # a maioria das consts do dashboard é um array ([...]); MUDANCAS_RECENTES
    # é o único objeto ({...}) atualizado por aqui, então o padrão se adapta
    # ao tipo do dado em vez de assumir colchetes.
    abre_fecha = (r"\{", r"\}") if isinstance(dados, dict) else (r"\[", r"\]")
    novo, n = re.subn(
        rf"const {nome} = {abre_fecha[0]}.*?{abre_fecha[1]};",
        f"const {nome} = {js(dados)};",
        html, count=1, flags=re.DOTALL,
    )
    if n == 0:
        raise SystemExit(f"ERRO: não encontrei 'const {nome}' no {HTML_PATH}.")
    if isinstance(dados, dict):
        log(f"  {nome}: atualizado")
    else:
        log(f"  {nome}: {len(dados)} itens")
    return novo


def ler_const(html, nome):
    """Extrai um array JS do HTML e devolve como lista Python."""
    m = re.search(rf"const {nome} = (\[.*?\]);", html, re.DOTALL)
    if not m:
        return []
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return []


# Mesma ordem de classificação já usada em outros dois lugares do projeto
# (a tabela de composição do e-mail de notificação, mais abaixo neste arquivo,
# e o array `ranks` usado pelo dashboard em `renderAprovacaoPanel`): do melhor
# para o pior. Não existe hoje uma constante central para isso — esta lista
# deriva exatamente da ordem já adotada nesses dois pontos, para não inventar
# um critério novo.
RANKING_ORDEM = ["A", "B", "BV", "C", "D", "X"]


def detectar_mudancas_ranking(lots_novos, lots_antigos):
    """Compara o ranking de cada lote entre o estado publicado antes desta
    execução (lots_antigos, capturado do index.html ANTES de ser sobrescrito)
    e o estado recém-calculado agora (lots_novos).

    Associa os registros por (lote, cultivar) — nunca por posição no array,
    já que a ordem pode mudar entre atualizações. Só entra no resultado
    quando o ranking realmente muda; lotes novos (sem registro anterior) e
    lotes que somem do dataset não contam como mudança de classificação.
    """
    posicao = {r: i for i, r in enumerate(RANKING_ORDEM)}
    ranking_anterior_por_chave = {
        (l.get("lote"), l.get("cultivar")): l.get("ranking")
        for l in lots_antigos if l.get("lote")
    }

    subiram, cairam = [], []
    for l in lots_novos:
        lote = l.get("lote")
        if not lote:
            continue
        chave = (lote, l.get("cultivar"))
        rank_anterior = ranking_anterior_por_chave.get(chave)
        rank_atual = l.get("ranking")

        if rank_anterior is None:
            continue  # lote novo — sem estado anterior, não é mudança de ranking
        if rank_anterior == rank_atual:
            continue  # classificação final igual, mesmo que germ/vigor oficiais tenham mudado
        if rank_anterior not in posicao or rank_atual not in posicao:
            continue  # ranking fora do vocabulário conhecido — ignora com segurança

        item = {"lote": lote, "cultivar": l.get("cultivar"), "de": rank_anterior, "para": rank_atual}
        if posicao[rank_atual] < posicao[rank_anterior]:
            subiram.append(item)
        else:
            cairam.append(item)

    return {
        "subiram": subiram,
        "cairam": cairam,
        "total_subiram": len(subiram),
        "total_cairam": len(cairam),
    }


def atualizar_html(lots, summary, pms_cv, comprados, benef):
    with open(HTML_PATH, encoding="utf-8") as f:
        html = f.read()

    # estado anterior — usado para montar o resumo do e-mail e para detectar
    # mudanças de ranking (MUDANCAS_RECENTES), sempre lido ANTES de qualquer
    # substituição neste html: é o que estava publicado antes desta execução.
    global ESTADO_ANTERIOR
    ESTADO_ANTERIOR = {
        "lots": ler_const(html, "LOTS"),
        "summary": ler_const(html, "SUMMARY"),
        "comprados": ler_const(html, "LOTES_COMPRADOS"),
        "benef": ler_const(html, "BENEF"),
    }

    # comparação (estado publicado) x (estado recém-calculado nesta execução) —
    # feita agora, antes de LOTS ser sobrescrito, para não comparar dados já
    # atualizados contra eles mesmos.
    mudancas_ranking = detectar_mudancas_ranking(lots, ESTADO_ANTERIOR["lots"])
    log(
        f"  mudanças de ranking: {mudancas_ranking['total_subiram']} subiram · "
        f"{mudancas_ranking['total_cairam']} caíram"
    )

    # guarda o SUMMARY atual como SUMMARY_PREV (alimenta o botão "Δ anterior")
    atual = re.search(r"const SUMMARY = (\[.*?\]);", html, re.DOTALL)
    if atual and "const SUMMARY_PREV" in html:
        html, _ = re.subn(
            r"const SUMMARY_PREV = \[.*?\];",
            f"const SUMMARY_PREV = {atual.group(1)};",
            html, count=1, flags=re.DOTALL,
        )

    html = substituir_const(html, "LOTS", lots)
    html = substituir_const(html, "MUDANCAS_RECENTES", mudancas_ranking)
    html = substituir_const(html, "SUMMARY", summary)
    html = substituir_const(html, "PMS_CV", pms_cv)
    html = substituir_const(html, "LOTES_COMPRADOS", comprados)
    if benef is None:
        log("  BENEF: preservado (planilha não informada)")
    else:
        html = substituir_const(html, "BENEF", benef)

    # diagnóstico da safra — narrativa gerada a partir dos dados desta execução
    # (antes era um texto estático, nunca atualizado pela automação)
    diag = montar_diagnostico(lots, summary)
    html = substituir_const(html, "DIAG_DATA", diag)

    # data no header e no rodapé
    html, _ = re.subn(
        r'(<div class="val" id="h-data"[^>]*>)[^<]*(</div>)',
        rf"\g<1>{CARIMBO_HEADER}\2", html,
    )
    html, _ = re.subn(
        r"Atualizado em [0-9/]+(?: \d{2}:\d{2})?",
        f"Atualizado em {CARIMBO_HEADER}", html,
    )

    # histórico de aprovação
    total = len(lots)
    aprov = sum(1 for l in lots if l["ranking"] in APROVADOS)
    cont = defaultdict(int)
    for l in lots:
        cont[l["ranking"]] += 1

    ponto = {
        "data": HOJE,
        "rank_A": cont["A"], "rank_B": cont["B"], "rank_BV": cont["BV"],
        "rank_C": cont["C"], "rank_D": cont["D"], "rank_X": cont["X"],
        "total": total,
        "pctAprov": round(aprov / total * 100, 1) if total else 0,
    }

    m = re.search(r"HISTORICO_APROVACAO = (\[.*?\]);", html, re.DOTALL)
    if m:
        hist = json.loads(m.group(1))
        chaves = ["rank_A", "rank_B", "rank_BV", "rank_C", "rank_D", "rank_X", "total"]

        def igual(p):
            return all(p.get(k) == ponto[k] for k in chaves)

        if any(igual(p) for p in hist):
            for p in hist:
                if igual(p):
                    p["data"] = HOJE
            log("  histórico: composição igual, data atualizada")
        else:
            hist.append(ponto)
            log(f"  histórico: novo ponto (#{len(hist)})")

        html, _ = re.subn(
            r"HISTORICO_APROVACAO = \[.*?\];",
            f"HISTORICO_APROVACAO = {js(hist)};",
            html, count=1, flags=re.DOTALL,
        )

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    return total, aprov, dict(cont)


# ---------------------------------------------------------------- resumo

ESTADO_ANTERIOR = {}

ROTULO_RANK = {
    "A": "Ótimo", "B": "Bom", "C": "Condicional",
    "BV": "Cond. Vigor", "D": "Ruim", "X": "Pendente",
}
COR_RANK = {
    "A": "#16A34A", "B": "#2563EB", "C": "#D97706",
    "BV": "#D97706", "D": "#DC2626", "X": "#64748B",
}


def contar_ranks(lista):
    c = defaultdict(int)
    for l in lista:
        c[l.get("ranking")] += 1
    return c


def montar_resumo(lots, summary, comprados, benef):
    """Compara o estado anterior com o novo e devolve (assunto, corpo_html, mudou)."""
    ant = ESTADO_ANTERIOR
    lots_ant = ant.get("lots", [])

    # --- números gerais ---
    total_novo, total_ant = len(lots), len(lots_ant)
    ap_novo = sum(1 for l in lots if l["ranking"] in APROVADOS)
    ap_ant  = sum(1 for l in lots_ant if l.get("ranking") in APROVADOS)
    pct_novo = round(ap_novo / total_novo * 100, 1) if total_novo else 0
    pct_ant  = round(ap_ant / total_ant * 100, 1) if total_ant else 0

    c_novo, c_ant = contar_ranks(lots), contar_ranks(lots_ant)

    # --- lotes novos e mudanças de classificação ---
    por_lote_ant = {l.get("lote"): l for l in lots_ant if l.get("lote")}
    novos, mudaram = [], []
    for l in lots:
        chave = l.get("lote")
        if not chave:
            continue
        anterior = por_lote_ant.get(chave)
        if anterior is None:
            novos.append(l)
        elif anterior.get("ranking") != l["ranking"]:
            mudaram.append((l, anterior.get("ranking")))

    # --- comprados e beneficiamento ---
    comp_ant = ant.get("comprados", [])
    bags_comp_novo = sum(c.get("bags") or 0 for c in comprados)
    bags_comp_ant  = sum(c.get("bags") or 0 for c in comp_ant)

    ben_ant = ant.get("benef", [])
    bags_ben_novo = sum(b.get("bags_total") or 0 for b in benef) if benef is not None else None
    bags_ben_ant  = sum(b.get("bags_total") or 0 for b in ben_ant)

    mudou = bool(
        novos or mudaram
        or total_novo != total_ant
        or len(comprados) != len(comp_ant)
        or bags_comp_novo != bags_comp_ant
        or (bags_ben_novo is not None and bags_ben_novo != bags_ben_ant)
    )

    # ---------------- montagem do HTML ----------------
    def seta(d, bom_subir=True):
        if d == 0:
            return '<span style="color:#94A3B8;">–</span>'
        cor = ("#16A34A" if d > 0 else "#DC2626") if bom_subir else ("#DC2626" if d > 0 else "#16A34A")
        return f'<span style="color:{cor};font-weight:700;">{"+" if d > 0 else ""}{d}</span>'

    d_pct = round(pct_novo - pct_ant, 1)
    cor_pct = "#16A34A" if d_pct > 0 else ("#DC2626" if d_pct < 0 else "#64748B")
    txt_pct = f'{"+" if d_pct > 0 else ""}{d_pct} pp' if d_pct else "sem variação"

    linhas_rank = ""
    for r in ["A", "B", "BV", "C", "D", "X"]:
        n, a = c_novo.get(r, 0), c_ant.get(r, 0)
        if n == 0 and a == 0:
            continue
        bom = r in APROVADOS
        linhas_rank += (
            f'<tr>'
            f'<td style="padding:7px 10px;border-bottom:1px solid #E2E8F0;">'
            f'<b style="color:{COR_RANK[r]};">{r}</b> '
            f'<span style="color:#64748B;font-size:12px;">{ROTULO_RANK[r]}</span></td>'
            f'<td style="padding:7px 10px;border-bottom:1px solid #E2E8F0;text-align:right;color:#64748B;">{a}</td>'
            f'<td style="padding:7px 10px;border-bottom:1px solid #E2E8F0;text-align:right;font-weight:700;">{n}</td>'
            f'<td style="padding:7px 10px;border-bottom:1px solid #E2E8F0;text-align:right;">{seta(n - a, bom)}</td>'
            f'</tr>'
        )

    blocos = ""

    if novos:
        porcv = defaultdict(int)
        for l in novos:
            porcv[l.get("cultivar") or "?"] += 1
        itens = " · ".join(f"{cv} ({q})" for cv, q in sorted(porcv.items(), key=lambda x: -x[1]))
        blocos += (
            f'<div style="margin-top:22px;">'
            f'<div style="font-size:13px;font-weight:700;color:#0F172A;margin-bottom:6px;">'
            f'{len(novos)} lote{"s" if len(novos) != 1 else ""} novo{"s" if len(novos) != 1 else ""}</div>'
            f'<div style="font-size:13px;color:#475569;line-height:1.6;">{itens}</div>'
            f'</div>'
        )

    if mudaram:
        linhas = ""
        for l, antes in mudaram[:20]:
            linhas += (
                f'<tr>'
                f'<td style="padding:5px 10px 5px 0;font-family:monospace;font-size:12px;">{l.get("lote")}</td>'
                f'<td style="padding:5px 10px 5px 0;font-size:12px;color:#475569;">{l.get("cultivar") or ""}</td>'
                f'<td style="padding:5px 0;font-size:12px;">'
                f'<span style="color:{COR_RANK.get(antes, "#64748B")};">{antes or "—"}</span>'
                f'<span style="color:#94A3B8;"> &rarr; </span>'
                f'<b style="color:{COR_RANK.get(l["ranking"], "#64748B")};">{l["ranking"]}</b>'
                f'</td></tr>'
            )
        resto = (
            f'<div style="font-size:12px;color:#94A3B8;margin-top:6px;">'
            f'e mais {len(mudaram) - 20} lote(s)</div>'
        ) if len(mudaram) > 20 else ""
        blocos += (
            f'<div style="margin-top:22px;">'
            f'<div style="font-size:13px;font-weight:700;color:#0F172A;margin-bottom:6px;">'
            f'{len(mudaram)} lote{"s" if len(mudaram) != 1 else ""} mudou de classificação</div>'
            f'<table style="border-collapse:collapse;">{linhas}</table>{resto}'
            f'</div>'
        )

    extras = ""
    if len(comprados) != len(comp_ant) or bags_comp_novo != bags_comp_ant:
        extras += (
            f'<li style="margin-bottom:4px;">Lotes comprados: <b>{len(comprados)}</b> '
            f'({seta(len(comprados) - len(comp_ant))}) · '
            f'<b>{bags_comp_novo:,}</b> bags ({seta(bags_comp_novo - bags_comp_ant)})</li>'
        ).replace(",", ".")
    if bags_ben_novo is not None and bags_ben_novo != bags_ben_ant:
        extras += (
            f'<li style="margin-bottom:4px;">Beneficiamento: <b>{bags_ben_novo:,}</b> bags '
            f'({seta(bags_ben_novo - bags_ben_ant)})</li>'
        ).replace(",", ".")
    if extras:
        blocos += (
            f'<div style="margin-top:22px;">'
            f'<div style="font-size:13px;font-weight:700;color:#0F172A;margin-bottom:6px;">Outras abas</div>'
            f'<ul style="margin:0;padding-left:18px;font-size:13px;color:#475569;">{extras}</ul>'
            f'</div>'
        )

    if not blocos:
        blocos = (
            '<div style="margin-top:22px;font-size:13px;color:#64748B;">'
            'Os totais permanecem iguais — apenas valores internos de análise foram revisados.'
            '</div>'
        )

    corpo = f"""<!DOCTYPE html>
<html><body style="margin:0;padding:24px;background:#F1F5F9;font-family:-apple-system,'Segoe UI',Roboto,Arial,sans-serif;">
<div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #E2E8F0;">

  <div style="background:#0F172A;padding:20px 24px;">
    <div style="color:#fff;font-size:17px;font-weight:800;">Gestão de Qualidade — Sementes com Vigor</div>
    <div style="color:#94A3B8;font-size:13px;margin-top:3px;">Dashboard atualizado em {HOJE}</div>
  </div>

  <div style="padding:24px;">

    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
      <tr>
        <td style="padding:12px;background:#F8FAFC;border-radius:8px;width:50%;">
          <div style="font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:.5px;">Taxa de aprovação</div>
          <div style="font-size:26px;font-weight:800;color:#0F172A;margin-top:2px;">{pct_novo}%</div>
          <div style="font-size:12px;color:{cor_pct};margin-top:2px;">{txt_pct}</div>
        </td>
        <td style="width:10px;"></td>
        <td style="padding:12px;background:#F8FAFC;border-radius:8px;width:50%;">
          <div style="font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:.5px;">Lotes analisados</div>
          <div style="font-size:26px;font-weight:800;color:#0F172A;margin-top:2px;">{total_novo}</div>
          <div style="font-size:12px;color:#64748B;margin-top:2px;">{ap_novo} aprovados</div>
        </td>
      </tr>
    </table>

    <div style="font-size:13px;font-weight:700;color:#0F172A;margin-bottom:8px;">Composição por classificação</div>
    <table style="width:100%;border-collapse:collapse;font-size:13px;">
      <tr style="background:#F8FAFC;">
        <th style="padding:7px 10px;text-align:left;font-size:11px;color:#64748B;font-weight:600;">Classificação</th>
        <th style="padding:7px 10px;text-align:right;font-size:11px;color:#64748B;font-weight:600;">Antes</th>
        <th style="padding:7px 10px;text-align:right;font-size:11px;color:#64748B;font-weight:600;">Agora</th>
        <th style="padding:7px 10px;text-align:right;font-size:11px;color:#64748B;font-weight:600;">Var.</th>
      </tr>
      {linhas_rank}
    </table>

    {blocos}

    <div style="margin-top:26px;text-align:center;">
      <a href="https://laisetimbira-hue.github.io/dashboard-soja-2526/"
         style="display:inline-block;background:#2563EB;color:#fff;text-decoration:none;
                padding:11px 26px;border-radius:8px;font-size:14px;font-weight:700;">
        Abrir o dashboard
      </a>
    </div>

  </div>

  <div style="padding:14px 24px;background:#F8FAFC;border-top:1px solid #E2E8F0;
              font-size:11px;color:#94A3B8;text-align:center;">
    Mensagem automática · gerada a partir da planilha no Google Drive
  </div>

</div></body></html>"""

    sinal = f"{'+' if d_pct > 0 else ''}{d_pct}pp" if d_pct else "estável"
    assunto = f"Dashboard atualizado — {pct_novo}% aprovação ({sinal}) · {total_novo} lotes · {HOJE}"

    return assunto, corpo, mudou


def gravar_resumo(assunto, corpo, mudou):
    with open("resumo_email.html", "w", encoding="utf-8") as f:
        f.write(corpo)
    saida = os.environ.get("GITHUB_OUTPUT")
    if saida:
        with open(saida, "a", encoding="utf-8") as f:
            f.write(f"assunto={assunto}\n")
            f.write(f"mudou={'true' if mudou else 'false'}\n")
    log(f"  resumo gerado · mudanças relevantes: {'sim' if mudou else 'não'}")


# ---------------------------------------------------------------- main

def main():
    log("Baixando planilhas do Drive...")
    buf_analises = baixar(ID_ANALISES, "ANALISES")
    buf_benef    = baixar(ID_BENEF, "BENEFICIADOS", obrigatorio=False)

    log("\nProcessando ANALISES...")
    lots = parse_analises(buf_analises)
    summary, pms_cv = montar_summary(lots)

    buf_analises.seek(0)
    comprados = parse_comprados(buf_analises)
    log(f"  {len(lots)} lotes · {len(comprados)} comprados")

    if buf_benef is None:
        benef = None
        log("\nBENEFICIADOS: pulado — o BENEF atual do index.html será preservado")
    else:
        log("\nProcessando BENEFICIADOS...")
        benef = parse_benef(buf_benef)
        log(f"  {len(benef)} entradas · {sum(b['bags_total'] for b in benef):,} bags")

    log(f"\nInjetando no {HTML_PATH}...")
    total, aprov, cont = atualizar_html(lots, summary, pms_cv, comprados, benef)

    log("\nMontando resumo para o e-mail...")
    assunto, corpo, mudou = montar_resumo(lots, summary, comprados, benef)
    gravar_resumo(assunto, corpo, mudou)

    pct = round(aprov / total * 100, 1) if total else 0
    log(f"\n{'='*46}")
    log(f"  {HOJE}")
    log(f"  {total} lotes · {aprov} aprovados ({pct}%)")
    log(f"  {cont}")
    log(f"{'='*46}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        log(f"\nFALHOU: {type(e).__name__}: {e}")
        sys.exit(1)
